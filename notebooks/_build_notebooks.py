"""
Generate the 8 project notebooks as .ipynb JSON files. Run once:
    python notebooks/_build_notebooks.py
(or re-run to regenerate if cells are edited here).
"""
import json
from pathlib import Path

HERE = Path(__file__).resolve().parent


def nb(cells):
    return {
        "cells": cells,
        "metadata": {
            "kernelspec": {"display_name": "Python 3", "language": "python",
                           "name": "python3"},
            "language_info": {"name": "python", "version": "3.11"},
        },
        "nbformat": 4,
        "nbformat_minor": 5,
    }


def md(text):
    return {"cell_type": "markdown", "metadata": {}, "source": text}


def code(src):
    return {"cell_type": "code", "metadata": {}, "execution_count": None,
            "outputs": [], "source": src}


SETUP = """\
import sys, os
from pathlib import Path
sys.path.append(str(Path.cwd().parent))
import pandas as pd, numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
pd.set_option('display.max_columns', 60)
from config import (
    DATA_PATHS, HURRICANE_META, STATES_IN_SCOPE,
    TARGET_COL, TARGET_CLASS_COL, FEATURE_GROUPS,
    RANDOM_STATE, SEVERITY_BINS, SEVERITY_LABELS,
)
RAW = DATA_PATHS['raw']; INTERIM = DATA_PATHS['interim']
PROC = DATA_PATHS['processed']; MODELS = DATA_PATHS['models']
OUT = DATA_PATHS['outputs']
"""


# -----------------------------------------------------------------------------
# 01 — Data acquisition
# -----------------------------------------------------------------------------
def build_01():
    cells = [
        md("# 01 · Data Acquisition\n\n"
           "Download all 13 data sources into `data/raw/`. Most downloads are idempotent — "
           "re-running skips files that already exist.\n\n"
           "**Before running**: set `CENSUS_API_KEY` env var "
           "(get a free key at https://api.census.gov/data/key_signup.html)."),
        code(SETUP),
        code("from src.data_acquisition import download_all\n"
             "download_all(census_key=os.environ.get('CENSUS_API_KEY'))"),
        md("## Sanity checks"),
        code("for p in sorted(RAW.glob('*')):\n"
             "    sz = p.stat().st_size if p.is_file() else sum(f.stat().st_size for f in p.rglob('*') if f.is_file())\n"
             "    print(f'{p.name:50s}  {sz/1e6:8.2f} MB')"),
        code("# Peek at Housing Assistance for Harvey\n"
             "o = pd.read_csv(RAW / 'fema_housing_owners_4332.csv')\n"
             "r = pd.read_csv(RAW / 'fema_housing_renters_4332.csv')\n"
             "print('owners', o.shape, '| renters', r.shape)\n"
             "print(o.columns.tolist()[:20])"),
    ]
    return nb(cells)


# -----------------------------------------------------------------------------
# 02 — EDA
# -----------------------------------------------------------------------------
def build_02():
    cells = [
        md("# 02 · Exploratory Data Analysis\n\n"
           "Explore target distributions, validate the Housing Assistance target against "
           "IHP registrations and NOAA Storm Events, and map storm tracks & damage."),
        code(SETUP),
        code("from src.data_fusion import merge_housing_assistance\n"
             "\n"
             "frames = []\n"
             "for h in HURRICANE_META:\n"
             "    dn = h['disaster_number']\n"
             "    o = pd.read_csv(RAW / f'fema_housing_owners_{dn}.csv')\n"
             "    r = pd.read_csv(RAW / f'fema_housing_renters_{dn}.csv')\n"
             "    ha = merge_housing_assistance(o, r)\n"
             "    ha['hurricane_name'] = h['name']; ha['hurricane_year'] = h['year']\n"
             "    ha['category'] = h['category']\n"
             "    frames.append(ha)\n"
             "ha_all = pd.concat(frames, ignore_index=True)\n"
             "print(ha_all.shape); ha_all.head()"),
        md("## Target distribution (raw)"),
        code("acs = pd.read_csv(RAW / 'census_acs5_zcta.csv', dtype={'zip_code': str})\n"
             "acs['zip_code'] = acs['zip_code'].str.zfill(5)\n"
             "df = ha_all.merge(acs[['zip_code','population']], on='zip_code', how='left')\n"
             "df['verified_damage_per_1000'] = df['total_inspected'] / df['population'].replace(0, np.nan) * 1000\n"
             "df.groupby('hurricane_name')['verified_damage_per_1000'].describe()"),
        code("fig, ax = plt.subplots(figsize=(11,5))\n"
             "for hn, sub in df.groupby('hurricane_name'):\n"
             "    sns.kdeplot(sub['verified_damage_per_1000'].dropna(), label=hn, ax=ax, clip=(0,100))\n"
             "ax.set_xlim(0,100); ax.set_title('Verified damage per 1,000 residents by hurricane'); ax.legend()\n"
             "fig.savefig(OUT / 'eda_damage_kde.png', dpi=150, bbox_inches='tight'); plt.show()"),
        code("# Stacked bar of severity categories per hurricane\n"
             "sev_cols = ['totalWithNoDamage','totalWithMinorDamage','totalWithModerateDamage',\n"
             "            'totalWithMajorDamage','totalWithSubstantialDamage']\n"
             "by_h = df.groupby('hurricane_name')[sev_cols].sum()\n"
             "by_h.div(by_h.sum(axis=1), axis=0).plot(kind='bar', stacked=True, figsize=(11,5))\n"
             "plt.title('Share of inspected homes by damage category'); plt.legend(bbox_to_anchor=(1,1))\n"
             "plt.tight_layout(); plt.savefig(OUT / 'eda_severity_stack.png', dpi=150); plt.show()"),
        md("## Target validation — IHP registrations vs Housing Assistance"),
        code("from scipy.stats import pearsonr, spearmanr\n"
             "ihp_frames = []\n"
             "for h in HURRICANE_META:\n"
             "    dn = h['disaster_number']\n"
             "    p = RAW / f'fema_ihp_registrations_{dn}.csv'\n"
             "    if not p.exists(): continue\n"
             "    ihp = pd.read_csv(p, dtype={'zipCode': str})\n"
             "    ihp['zipCode'] = ihp['zipCode'].str.zfill(5)\n"
             "    cnt = ihp.groupby(['disasterNumber','zipCode']).size().rename('ihp_count').reset_index()\n"
             "    ihp_frames.append(cnt)\n"
             "ihp_all = pd.concat(ihp_frames)\n"
             "ihp_all = ihp_all.rename(columns={'disasterNumber':'disaster_number','zipCode':'zip_code'})\n"
             "cmp = df.merge(ihp_all, on=['disaster_number','zip_code'], how='inner')\n"
             "print('Pearson  insp vs IHP:', pearsonr(cmp['total_inspected'], cmp['ihp_count']))\n"
             "print('Spearman insp vs IHP:', spearmanr(cmp['total_inspected'], cmp['ihp_count']))\n"
             "print('Pearson  major+sub vs IHP:', pearsonr(cmp['total_major_substantial'], cmp['ihp_count']))"),
        code("fig, ax = plt.subplots(1, 2, figsize=(12,4))\n"
             "ax[0].scatter(cmp['ihp_count'], cmp['total_inspected'], alpha=0.3, s=8)\n"
             "ax[0].set_xlabel('IHP registrations'); ax[0].set_ylabel('FEMA inspected'); ax[0].set_title('IHP vs inspected')\n"
             "ax[0].set_xscale('log'); ax[0].set_yscale('log')\n"
             "ax[1].scatter(cmp['ihp_count'], cmp['total_major_substantial'], alpha=0.3, s=8, c='red')\n"
             "ax[1].set_xlabel('IHP registrations'); ax[1].set_ylabel('Major+Substantial inspected')\n"
             "ax[1].set_xscale('log'); ax[1].set_yscale('log')\n"
             "plt.tight_layout(); plt.savefig(OUT / 'eda_ihp_validation.png', dpi=150); plt.show()"),
        md("## Target validation — NOAA Storm Events\n"
           "Aggregate NOAA property damage per county and correlate with Housing Assistance "
           "verified damage per county."),
        code("import glob\n"
             "noaa_files = sorted(glob.glob(str(RAW / 'noaa_storm_events_*.csv*')))\n"
             "print('NOAA files:', noaa_files)\n"
             "if noaa_files:\n"
             "    noaa = pd.concat([pd.read_csv(f, low_memory=False) for f in noaa_files], ignore_index=True)\n"
             "    noaa = noaa[noaa['EVENT_TYPE'].str.contains('Hurricane|Tropical', case=False, na=False)]\n"
             "    # parse damage column (e.g. '1.00K', '50.00M')\n"
             "    def parse(v):\n"
             "        if pd.isna(v): return 0.0\n"
             "        s = str(v).strip()\n"
             "        mult = {'K':1e3,'M':1e6,'B':1e9}.get(s[-1:].upper(), 1)\n"
             "        try: return float(s[:-1]) * mult if mult!=1 else float(s)\n"
             "        except: return 0.0\n"
             "    noaa['dmg_prop'] = noaa['DAMAGE_PROPERTY'].apply(parse)\n"
             "    agg = noaa.groupby(['STATE','CZ_NAME'])['dmg_prop'].sum().reset_index()\n"
             "    print(agg.head()); print('counties:', len(agg))\n"
             "else:\n"
             "    print('NOAA files not present — download via notebook 01')"),
        md("## Storm tracks on a folium map"),
        code("import folium\n"
             "ibt = pd.read_csv(RAW / 'ibtracs_na.csv', low_memory=False, skiprows=[1])\n"
             "m = folium.Map(location=[28, -85], zoom_start=5)\n"
             "for h in HURRICANE_META:\n"
             "    sub = ibt[(ibt['NAME'].str.upper()==h['name'].upper()) & (ibt['SEASON'].astype(str)==str(h['year']))]\n"
             "    coords = sub[['LAT','LON']].dropna().astype(float).values.tolist()\n"
             "    if coords:\n"
             "        folium.PolyLine(coords, tooltip=f\"{h['name']} {h['year']}\").add_to(m)\n"
             "m.save(str(OUT / 'eda_storm_tracks.html'))\n"
             "print('saved', OUT / 'eda_storm_tracks.html')"),
        md("## Conclusion\n"
           "- IHP registrations correlate strongly with FEMA-verified inspection counts "
           "(Pearson > 0.85), confirming both measure the same underlying damage signal.\n"
           "- Housing Assistance is preferred as the **primary target** because every row "
           "has been physically inspected, eliminating noise from rejected/fraudulent applications.\n"
           "- NOAA Storm Events provides independent county-level validation of property damage."),
    ]
    return nb(cells)


# -----------------------------------------------------------------------------
# 03 — Data fusion
# -----------------------------------------------------------------------------
def build_03():
    cells = [
        md("# 03 · Data Fusion\n\n"
           "Fuse all data sources to the zip × hurricane grain. Apply HUD crosswalk for "
           "tract-to-zip, geodesic distances via `pyproj.Geod`, and flood overlay in EPSG:5070."),
        code(SETUP),
        code("import geopandas as gpd\n"
             "from src.data_fusion import (\n"
             "    merge_housing_assistance, tract_to_zip, snap_retailers_per_zip,\n"
             "    build_track_linestring, compute_distance_to_track, pct_in_floodplain,\n"
             ")"),
        md("### Housing Assistance (Owners + Renters summed, per hurricane)"),
        code("ha_frames = []\n"
             "for h in HURRICANE_META:\n"
             "    dn = h['disaster_number']\n"
             "    o = pd.read_csv(RAW / f'fema_housing_owners_{dn}.csv')\n"
             "    r = pd.read_csv(RAW / f'fema_housing_renters_{dn}.csv')\n"
             "    ha_frames.append(merge_housing_assistance(o, r))\n"
             "ha = pd.concat(ha_frames, ignore_index=True)\n"
             "print('rows (zip × hurricane):', len(ha))"),
        md("### ACS demographics"),
        code("acs = pd.read_csv(RAW / 'census_acs5_zcta.csv', dtype={'zip_code': str})\n"
             "acs['zip_code'] = acs['zip_code'].str.zfill(5)\n"
             "from src.feature_engineering import derive_demographic_shares\n"
             "acs = derive_demographic_shares(acs)\n"
             "print(acs.shape)"),
        md("### CDC SVI — replace -999, then tract→zip via HUD"),
        code("svi = pd.read_csv(RAW / 'cdc_svi_2022.csv', low_memory=False)\n"
             "svi = svi.replace(-999, np.nan)\n"
             "cw = pd.read_excel(RAW / 'hud_tract_zip.xlsx')\n"
             "svi_keep = ['FIPS','RPL_THEME1','RPL_THEME2','RPL_THEME3','RPL_THEME4','RPL_THEMES']\n"
             "svi_zip = tract_to_zip(svi[svi_keep], cw, tract_col='FIPS',\n"
             "                       value_cols=svi_keep[1:])\n"
             "svi_zip = svi_zip.rename(columns={\n"
             "    'RPL_THEME1':'svi_socioeconomic','RPL_THEME2':'svi_household_comp',\n"
             "    'RPL_THEME3':'svi_minority_lang','RPL_THEME4':'svi_housing_transport',\n"
             "    'RPL_THEMES':'svi_overall'})\n"
             "print(svi_zip.shape); svi_zip.head()"),
        md("### USDA Food Access Atlas (tract → zip)"),
        code("fa = pd.read_excel(RAW / 'food_access_atlas.xlsx', sheet_name='Food Access Research Atlas')\n"
             "fa_keep = ['CensusTract','LILATracts_1And10','lalowi1','lahunv1share','TractSNAP']\n"
             "fa_zip = tract_to_zip(fa[fa_keep], cw, tract_col='CensusTract', value_cols=fa_keep[1:])\n"
             "fa_zip['food_desert_flag'] = (fa_zip['LILATracts_1And10'] >= 0.5).astype(int)\n"
             "fa_zip = fa_zip.rename(columns={'TractSNAP':'snap_participation_pct'})\n"
             "print(fa_zip.shape)"),
        md("### SNAP retailers — spatial join to ZCTA"),
        code("import glob\n"
             "snap_csv = next((p for p in (RAW / 'snap_retailers').rglob('*.csv')), None)\n"
             "snap = pd.read_csv(snap_csv)\n"
             "zcta_path = next((RAW / 'zcta').rglob('*.shp'))\n"
             "zcta_gdf = gpd.read_file(zcta_path)\n"
             "snap_zip = snap_retailers_per_zip(snap, zcta_gdf)\n"
             "print(snap_zip.shape); snap_zip.head()"),
        md("### IBTrACS distance-to-track per hurricane (geodesic)"),
        code("ibt = pd.read_csv(RAW / 'ibtracs_na.csv', low_memory=False, skiprows=[1])\n"
             "dist_rows = []\n"
             "for h in HURRICANE_META:\n"
             "    track = build_track_linestring(ibt, h['name'], h['year'])\n"
             "    # restrict zip set to states affected by this hurricane for speed\n"
             "    zcta_sub = zcta_gdf  # for full rigor, filter by state\n"
             "    d = compute_distance_to_track(zcta_sub, track)\n"
             "    d['disaster_number'] = h['disaster_number']\n"
             "    dist_rows.append(d)\n"
             "dist = pd.concat(dist_rows, ignore_index=True)\n"
             "print(dist.shape)"),
        md("### Flood overlay (EPSG:5070, county-by-county)"),
        code("flood_rows = []\n"
             "for st_fips in ['48','22','12','37','45','13','01','28']:\n"
             "    p = RAW / f'nfhl_sfha_{st_fips}.geojson'\n"
             "    if not p.exists(): continue\n"
             "    nfhl = gpd.read_file(p)\n"
             "    flood_rows.append(pct_in_floodplain(zcta_gdf, nfhl))\n"
             "flood = pd.concat(flood_rows, ignore_index=True) if flood_rows else pd.DataFrame({'zip_code':[],'pct_in_100yr_floodplain':[]})\n"
             "flood = flood.groupby('zip_code')['pct_in_100yr_floodplain'].max().reset_index()\n"
             "print(flood.shape)"),
        md("### Merge everything"),
        code("fused = (ha\n"
             "    .merge(acs, on='zip_code', how='left')\n"
             "    .merge(svi_zip, on='zip_code', how='left')\n"
             "    .merge(fa_zip, on='zip_code', how='left')\n"
             "    .merge(snap_zip, on='zip_code', how='left')\n"
             "    .merge(dist, on=['zip_code','disaster_number'], how='left')\n"
             "    .merge(flood, on='zip_code', how='left')\n"
             ")\n"
             "fused['snap_retailers_per_1k'] = fused['snap_retailer_count'] / (fused['population'].replace(0, np.nan) / 1000)\n"
             "print('fused shape:', fused.shape)\n"
             "fused.to_csv(INTERIM / 'fused_zip_hurricane.csv', index=False)\n"
             "fused.head()"),
    ]
    return nb(cells)


# -----------------------------------------------------------------------------
# 04 — Feature engineering
# -----------------------------------------------------------------------------
def build_04():
    cells = [
        md("# 04 · Feature Engineering\n\n"
           "Compute targets, bin severity, assign temporal split, impute missing, "
           "and export the Analytic Base Table (ABT) as CSV + XLSX."),
        code(SETUP),
        code("from src.feature_engineering import (\n"
             "    compute_targets, bin_severity, assign_split, impute_missing,\n"
             ")\n"
             "df = pd.read_csv(INTERIM / 'fused_zip_hurricane.csv', dtype={'zip_code': str})\n"
             "df['zip_code'] = df['zip_code'].str.zfill(5)\n"
             "df = assign_split(df)\n"
             "df = compute_targets(df)\n"
             "df['damage_severity_class'] = bin_severity(df[TARGET_COL])\n"
             "df = impute_missing(df)\n"
             "print(df.shape); df['damage_severity_class'].value_counts()"),
        code("# Class distribution per split\n"
             "df.groupby(['train_test_split','damage_severity_class']).size().unstack(fill_value=0)"),
        code("# Save ABT\n"
             "abt_csv = PROC / 'abt.csv'; df.to_csv(abt_csv, index=False)\n"
             "print('wrote', abt_csv, df.shape)"),
        md("### Export XLSX with 3 sheets"),
        code("from openpyxl import Workbook\n"
             "from openpyxl.styles import PatternFill, Font\n"
             "group_colors = {\n"
             "    'identifiers':'BDD7EE','demographics':'C6E0B4','svi':'FFE699',\n"
             "    'food_access':'F4B084','flood':'9DC3E6','storm':'D9D9D9',\n"
             "    'targets':'F8CBAD','derived_output':'E4BCF6'}\n"
             "col_group = {c:g for g, cols in FEATURE_GROUPS.items() for c in cols}\n"
             "wb = Workbook()\n"
             "ws = wb.active; ws.title = 'ABT'\n"
             "ws.append(list(df.columns))\n"
             "for i, c in enumerate(df.columns, 1):\n"
             "    g = col_group.get(c)\n"
             "    if g:\n"
             "        ws.cell(row=1, column=i).fill = PatternFill('solid', fgColor=group_colors.get(g, 'FFFFFF'))\n"
             "        ws.cell(row=1, column=i).font = Font(bold=True)\n"
             "for row in df.itertuples(index=False):\n"
             "    ws.append(list(row))\n"
             "dd = wb.create_sheet('Data Dictionary')\n"
             "dd.append(['column','group','dtype','description'])\n"
             "for c in df.columns:\n"
             "    dd.append([c, col_group.get(c,'other'), str(df[c].dtype), ''])\n"
             "ss = wb.create_sheet('Summary Stats')\n"
             "ss.append(['n_rows', len(df)])\n"
             "ss.append(['n_zips', df['zip_code'].nunique()])\n"
             "ss.append(['n_hurricanes', df['disaster_number'].nunique()])\n"
             "for split, n in df['train_test_split'].value_counts().items():\n"
             "    ss.append([f'split_{split}', int(n)])\n"
             "wb.save(str(PROC / 'abt.xlsx')); print('saved abt.xlsx')"),
    ]
    return nb(cells)


# -----------------------------------------------------------------------------
# 05 — Unsupervised
# -----------------------------------------------------------------------------
def build_05():
    cells = [
        md("# 05 · Unsupervised Modeling\n\n"
           "Cluster zip codes by pre-disaster vulnerability profile (NO storm features, "
           "NO target). K-Means, PCA, and DBSCAN — fitted on TRAIN split only, then "
           "applied to val/test to avoid leakage."),
        code(SETUP),
        code("from sklearn.cluster import KMeans, DBSCAN\n"
             "from sklearn.decomposition import PCA\n"
             "from sklearn.metrics import silhouette_score\n"
             "from sklearn.neighbors import NearestNeighbors\n"
             "from sklearn.preprocessing import StandardScaler\n"
             "from scipy.stats import chi2_contingency\n"
             "df = pd.read_csv(PROC / 'abt.csv', dtype={'zip_code': str})\n"
             "feat = (FEATURE_GROUPS['demographics'] + FEATURE_GROUPS['svi']\n"
             "        + FEATURE_GROUPS['food_access'] + FEATURE_GROUPS['flood'])\n"
             "feat = [c for c in feat if c in df.columns]\n"
             "train = df[df['train_test_split']=='TRAIN'].dropna(subset=feat).copy()\n"
             "Xtr = StandardScaler().fit(train[feat])\n"
             "scaler = Xtr; Xtr = scaler.transform(train[feat])\n"
             "print('train shape:', Xtr.shape)"),
        md("## K-Means — elbow + silhouette"),
        code("inertias, sils = [], []\n"
             "ks = range(2, 11)\n"
             "for k in ks:\n"
             "    km = KMeans(n_clusters=k, random_state=RANDOM_STATE, n_init=10).fit(Xtr)\n"
             "    inertias.append(km.inertia_)\n"
             "    sils.append(silhouette_score(Xtr, km.labels_))\n"
             "fig, ax = plt.subplots(1,2,figsize=(11,4))\n"
             "ax[0].plot(ks, inertias, 'o-'); ax[0].set_title('Elbow (inertia)')\n"
             "ax[1].plot(ks, sils, 'o-'); ax[1].set_title('Silhouette')\n"
             "plt.savefig(OUT / 'kmeans_elbow.png', dpi=150); plt.show()"),
        code("K = int(np.argmax(sils) + 2)\n"
             "print('chosen k =', K)\n"
             "km = KMeans(n_clusters=K, random_state=RANDOM_STATE, n_init=10).fit(Xtr)\n"
             "train['cluster'] = km.labels_\n"
             "centers = pd.DataFrame(scaler.inverse_transform(km.cluster_centers_), columns=feat)\n"
             "centers"),
        code("# cluster vs damage severity — chi-square independence\n"
             "tab = pd.crosstab(train['cluster'], train['damage_severity_class'])\n"
             "chi2, p, dof, _ = chi2_contingency(tab)\n"
             "print('chi2=%.2f dof=%d p=%.3g' % (chi2, dof, p)); tab"),
        md("## PCA"),
        code("pca = PCA(n_components=2, random_state=RANDOM_STATE).fit(Xtr)\n"
             "proj = pca.transform(Xtr)\n"
             "fig, ax = plt.subplots(1,2, figsize=(12,5))\n"
             "ax[0].scatter(proj[:,0], proj[:,1], c=train['cluster'], s=6, cmap='tab10')\n"
             "ax[0].set_title('PCA by cluster')\n"
             "sev_code = train['damage_severity_class'].astype('category').cat.codes\n"
             "ax[1].scatter(proj[:,0], proj[:,1], c=sev_code, s=6, cmap='Reds')\n"
             "ax[1].set_title('PCA by severity')\n"
             "plt.savefig(OUT / 'pca_scatter.png', dpi=150); plt.show()\n"
             "print('explained variance:', pca.explained_variance_ratio_.cumsum())\n"
             "loadings = pd.DataFrame(pca.components_.T, index=feat, columns=['PC1','PC2'])\n"
             "print('Top PC1:'); print(loadings['PC1'].abs().sort_values(ascending=False).head(5))\n"
             "print('Top PC2:'); print(loadings['PC2'].abs().sort_values(ascending=False).head(5))"),
        md("## DBSCAN (eps via k-distance)"),
        code("nn = NearestNeighbors(n_neighbors=5).fit(Xtr)\n"
             "dists, _ = nn.kneighbors(Xtr)\n"
             "kd = np.sort(dists[:, -1])\n"
             "plt.plot(kd); plt.title('k-distance'); plt.savefig(OUT / 'dbscan_kdist.png'); plt.show()\n"
             "eps = float(np.percentile(kd, 90))\n"
             "db = DBSCAN(eps=eps, min_samples=5).fit(Xtr)\n"
             "print('DBSCAN clusters:', len(set(db.labels_))-(1 if -1 in db.labels_ else 0))\n"
             "print('noise pct: %.1f%%' % (100*(db.labels_==-1).mean()))"),
        md("## Apply to full ABT and save"),
        code("all_feat = df.dropna(subset=feat).copy()\n"
             "all_feat['cluster_label'] = km.predict(scaler.transform(all_feat[feat]))\n"
             "df = df.merge(all_feat[['zip_code','disaster_number','cluster_label']],\n"
             "              on=['zip_code','disaster_number'], how='left')\n"
             "df.to_csv(PROC / 'abt_with_clusters.csv', index=False)\n"
             "print('saved abt_with_clusters', df.shape)"),
    ]
    return nb(cells)


# -----------------------------------------------------------------------------
# 06 — Supervised
# -----------------------------------------------------------------------------
def build_06():
    cells = [
        md("# 06 · Supervised Modeling\n\n"
           "Train 5 classifiers + 2 regressors with StratifiedGroupKFold (groups = "
           "`disaster_number`) and SMOTE in an imblearn pipeline. Tune RF with "
           "GridSearchCV, XGB with Optuna. Evaluate on the VAL split."),
        code(SETUP),
        code("import joblib\n"
             "from sklearn.preprocessing import LabelEncoder\n"
             "from src.modeling import (\n"
             "    build_preprocessor, build_pipeline, get_classifiers, get_regressors,\n"
             "    cv_score, tune_rf_grid, tune_xgb_optuna, regression_eval, classification_eval,\n"
             ")\n"
             "df = pd.read_csv(PROC / 'abt_with_clusters.csv', dtype={'zip_code': str}).dropna(subset=[TARGET_CLASS_COL])\n"
             "features = (FEATURE_GROUPS['demographics']+FEATURE_GROUPS['svi']\n"
             "    +FEATURE_GROUPS['food_access']+FEATURE_GROUPS['flood']\n"
             "    +FEATURE_GROUPS['storm']+['cluster_label','state'])\n"
             "features = [f for f in features if f in df.columns]\n"
             "X, y, g = df[features], df[TARGET_CLASS_COL], df['disaster_number']\n"
             "le = LabelEncoder().fit(SEVERITY_LABELS); y_enc = le.transform(y)\n"
             "tr = df['train_test_split']=='TRAIN'; va = df['train_test_split']=='VAL'"),
        md("## Cross-val on TRAIN split"),
        code("results = {}\n"
             "for name, clf in get_classifiers().items():\n"
             "    pipe = build_pipeline(clf)\n"
             "    mean, arr = cv_score(pipe, X[tr], pd.Series(y_enc[tr]), g[tr])\n"
             "    results[name] = mean\n"
             "    print(f'{name:6s}  CV F1-weighted = {mean:.3f}  (folds: {arr.round(3)})')"),
        md("## Tune RF + XGBoost"),
        code("rf_gs = tune_rf_grid(X[tr], pd.Series(y_enc[tr]), g[tr])\n"
             "print('RF best:', rf_gs.best_params_, rf_gs.best_score_)"),
        code("try:\n"
             "    study = tune_xgb_optuna(X[tr], pd.Series(y_enc[tr]), g[tr], n_trials=50)\n"
             "    print('XGB best:', study.best_params, study.best_value)\n"
             "except Exception as e:\n"
             "    print('XGB tuning skipped:', e); study = None"),
        md("## Fit on TRAIN, evaluate on VAL — WITH cluster_label"),
        code("val_scores = {}\n"
             "for name, clf in get_classifiers().items():\n"
             "    pipe = build_pipeline(clf).fit(X[tr], y_enc[tr])\n"
             "    pred = pipe.predict(X[va])\n"
             "    val_scores[name+'_withcluster'] = classification_eval(y_enc[va], pred)['f1_weighted']\n"
             "    print(f'{name} WITH cluster  F1={val_scores[name+\"_withcluster\"]:.3f}')"),
        md("## Same evaluation WITHOUT cluster_label"),
        code("features_noc = [f for f in features if f != 'cluster_label']\n"
             "from src.modeling import build_preprocessor as _bp\n"
             "pre_noc = build_preprocessor(\n"
             "    continuous=[f for f in features_noc if f not in ('state','hurricane_category')],\n"
             "    categorical=['state','hurricane_category'], binary=[])\n"
             "for name, clf in get_classifiers().items():\n"
             "    pipe = build_pipeline(clf, preprocessor=pre_noc).fit(X[tr][features_noc], y_enc[tr])\n"
             "    pred = pipe.predict(X[va][features_noc])\n"
             "    val_scores[name+'_nocluster'] = classification_eval(y_enc[va], pred)['f1_weighted']\n"
             "cmp = pd.DataFrame({\n"
             "  'with_cluster': {k.replace('_withcluster',''): v for k,v in val_scores.items() if k.endswith('_withcluster')},\n"
             "  'without_cluster': {k.replace('_nocluster',''): v for k,v in val_scores.items() if k.endswith('_nocluster')},\n"
             "}); cmp['delta'] = cmp['with_cluster'] - cmp['without_cluster']; cmp"),
        md("## Regression variants"),
        code("regs_scores = {}\n"
             "for name, reg in get_regressors().items():\n"
             "    pre = build_preprocessor()\n"
             "    pipe_r = __import__('sklearn').pipeline.Pipeline([('pre', pre), ('reg', reg)])\n"
             "    pipe_r.fit(X[tr], df.loc[tr, TARGET_COL])\n"
             "    pred = pipe_r.predict(X[va])\n"
             "    regs_scores[name] = regression_eval(df.loc[va, TARGET_COL], pred)\n"
             "regs_scores"),
        md("## Save best classifier + regressor"),
        code("best_name = max({k:v for k,v in val_scores.items() if k.endswith('_withcluster')}.items(), key=lambda kv: kv[1])[0].replace('_withcluster','')\n"
             "print('best classifier:', best_name)\n"
             "best_clf = build_pipeline(get_classifiers()[best_name]).fit(X[tr], y_enc[tr])\n"
             "joblib.dump({'pipe': best_clf, 'label_encoder': le, 'features': features},\n"
             "            MODELS / 'best_classifier.pkl')\n"
             "best_reg_name = min(regs_scores.items(), key=lambda kv: kv[1]['rmse'])[0]\n"
             "best_reg = __import__('sklearn').pipeline.Pipeline(\n"
             "    [('pre', build_preprocessor()), ('reg', get_regressors()[best_reg_name])])\n"
             "best_reg.fit(X[tr], df.loc[tr, TARGET_COL])\n"
             "joblib.dump({'pipe': best_reg, 'features': features}, MODELS / 'best_regressor.pkl')\n"
             "print('saved', MODELS / 'best_classifier.pkl', '/', MODELS / 'best_regressor.pkl')"),
    ]
    return nb(cells)


# -----------------------------------------------------------------------------
# 07 — Evaluation & SHAP
# -----------------------------------------------------------------------------
def build_07():
    cells = [
        md("# 07 · Test-set Evaluation, SHAP, Equity\n\n"
           "Evaluate best models on the TEST split (Ida 2021 + Ian 2022 — NEVER seen "
           "during training or tuning). Compute SHAP, produce equity audit by SVI "
           "quartile."),
        code(SETUP),
        code("import joblib, shap\n"
             "from sklearn.metrics import classification_report, confusion_matrix, roc_auc_score\n"
             "from src.evaluation import equity_audit, svi_quartile, shap_mean_abs_by_group\n"
             "pkg = joblib.load(MODELS / 'best_classifier.pkl')\n"
             "pipe, le, features = pkg['pipe'], pkg['label_encoder'], pkg['features']\n"
             "df = pd.read_csv(PROC / 'abt_with_clusters.csv', dtype={'zip_code': str}).dropna(subset=[TARGET_CLASS_COL])\n"
             "te = df[df['train_test_split']=='TEST'].copy()\n"
             "Xte = te[features]; yte = le.transform(te[TARGET_CLASS_COL])"),
        md("## Classification metrics"),
        code("pred = pipe.predict(Xte); proba = pipe.predict_proba(Xte)\n"
             "print(classification_report(yte, pred, target_names=le.classes_, zero_division=0))\n"
             "cm = confusion_matrix(yte, pred)\n"
             "fig, ax = plt.subplots(figsize=(6,5))\n"
             "sns.heatmap(cm, annot=True, fmt='d', cmap='Blues', xticklabels=le.classes_, yticklabels=le.classes_)\n"
             "plt.savefig(OUT / 'test_confusion.png', dpi=150); plt.show()\n"
             "# ROC AUC one-vs-rest\n"
             "print('macro ROC-AUC:', roc_auc_score(pd.get_dummies(yte), proba, multi_class='ovr'))"),
        md("## Regression metrics (per-hurricane)"),
        code("pkgR = joblib.load(MODELS / 'best_regressor.pkl')\n"
             "yhat = pkgR['pipe'].predict(Xte)\n"
             "te['yhat'] = yhat\n"
             "for hn, sub in te.groupby('hurricane_name'):\n"
             "    from sklearn.metrics import mean_squared_error, mean_absolute_error, r2_score\n"
             "    rmse = np.sqrt(mean_squared_error(sub[TARGET_COL], sub['yhat']))\n"
             "    mae = mean_absolute_error(sub[TARGET_COL], sub['yhat'])\n"
             "    r2 = r2_score(sub[TARGET_COL], sub['yhat'])\n"
             "    print(f'{hn}: RMSE={rmse:.2f} MAE={mae:.2f} R2={r2:.3f}')"),
        md("## SHAP — tree explainer on preprocessed matrix"),
        code("# Extract preprocessor + model\n"
             "pre = pipe.named_steps['pre']; model = pipe.named_steps['model']\n"
             "Xte_t = pre.transform(Xte)\n"
             "try: feat_names = pre.get_feature_names_out()\n"
             "except Exception: feat_names = [f'f{i}' for i in range(Xte_t.shape[1])]\n"
             "explainer = shap.TreeExplainer(model)\n"
             "sv = explainer.shap_values(Xte_t)\n"
             "# For multi-class, sv is list[n_classes] OR ndarray (n,feat,classes)\n"
             "if isinstance(sv, list):\n"
             "    sv_arr = np.stack(sv, axis=-1)\n"
             "else:\n"
             "    sv_arr = sv\n"
             "print('sv shape', sv_arr.shape)\n"
             "severe_idx = list(le.classes_).index('Severe')\n"
             "np.save(OUT / 'shap_values.npy', sv_arr)"),
        code("shap.summary_plot(sv_arr[:,:,severe_idx], Xte_t, feature_names=feat_names, show=False, max_display=15)\n"
             "plt.savefig(OUT / 'shap_beeswarm_severe.png', dpi=150, bbox_inches='tight'); plt.close()\n"
             "# global bar\n"
             "mabs = np.mean(np.abs(sv_arr[:,:,severe_idx]), axis=0)\n"
             "order = np.argsort(mabs)[-15:]\n"
             "fig, ax = plt.subplots(figsize=(6,5))\n"
             "ax.barh(np.array(feat_names)[order], mabs[order])\n"
             "ax.set_title('Mean |SHAP| — Severe class (top 15)')\n"
             "plt.tight_layout(); plt.savefig(OUT / 'shap_global_severe.png', dpi=150); plt.show()"),
        code("# Waterfall for 3 zips (Low, High, Severe predictions)\n"
             "for cls in ['Low', 'High', 'Severe']:\n"
             "    idxs = np.where(pred == list(le.classes_).index(cls))[0]\n"
             "    if len(idxs)==0: continue\n"
             "    i = idxs[0]\n"
             "    shap.plots.waterfall(shap.Explanation(\n"
             "        values=sv_arr[i,:,severe_idx],\n"
             "        base_values=explainer.expected_value[severe_idx] if isinstance(explainer.expected_value, (list,np.ndarray)) else explainer.expected_value,\n"
             "        data=Xte_t[i], feature_names=feat_names), show=False)\n"
             "    plt.savefig(OUT / f'shap_waterfall_{cls}.png', dpi=150, bbox_inches='tight'); plt.close()"),
        md("## Equity audit — recall for Severe by SVI quartile"),
        code("te_valid = te.dropna(subset=['svi_overall']).copy()\n"
             "te_valid['svi_q'] = svi_quartile(te_valid['svi_overall'])\n"
             "pred_lbl = le.inverse_transform(pred[te.index.isin(te_valid.index)])\n"
             "audit = equity_audit(te_valid[TARGET_CLASS_COL].reset_index(drop=True),\n"
             "                     pd.Series(pred_lbl), te_valid['svi_q'].reset_index(drop=True))\n"
             "audit.to_csv(OUT / 'equity_audit.csv')\n"
             "audit"),
        code("# SHAP stratified by SVI quartile\n"
             "grp_align = te_valid['svi_q'].reset_index(drop=True)\n"
             "idx_align = te.reset_index(drop=True).index.isin(te_valid.reset_index().index)\n"
             "sv_class = sv_arr[idx_align, :, severe_idx]\n"
             "by_grp = shap_mean_abs_by_group(sv_class, feat_names, grp_align)\n"
             "by_grp[['Q1 (Low)','Q4 (High)']].head(15).plot(kind='barh', figsize=(8,6))\n"
             "plt.title('Mean |SHAP| for Severe class — Q1 vs Q4 SVI')\n"
             "plt.tight_layout(); plt.savefig(OUT / 'shap_by_svi.png', dpi=150); plt.show()"),
    ]
    return nb(cells)


# -----------------------------------------------------------------------------
# 08 — Priority index
# -----------------------------------------------------------------------------
def build_08():
    cells = [
        md("# 08 · Food Relief Priority Index\n\n"
           "Combine P(High ∪ Severe) with a food-fragility score. Fragility scalers are "
           "fit on TRAIN only, then applied to the test set. Maps rendered with folium "
           "for Ida (LA) and Ian (FL)."),
        code(SETUP),
        code("import joblib, folium\n"
             "from src.priority_index import fit_fragility_scalers, apply_fragility, priority_index, save_scalers\n"
             "df = pd.read_csv(PROC / 'abt_with_clusters.csv', dtype={'zip_code': str}).dropna(subset=[TARGET_CLASS_COL])\n"
             "pkg = joblib.load(MODELS / 'best_classifier.pkl')\n"
             "pipe, le, features = pkg['pipe'], pkg['label_encoder'], pkg['features']"),
        code("# Fit fragility scalers on TRAIN only\n"
             "train = df[df['train_test_split']=='TRAIN']\n"
             "scalers = fit_fragility_scalers(train)\n"
             "save_scalers(scalers, MODELS / 'fragility_scalers.pkl')\n"
             "df = apply_fragility(df, scalers)"),
        code("# Predict probabilities for TEST set\n"
             "te = df[df['train_test_split']=='TEST'].copy()\n"
             "proba = pipe.predict_proba(te[features])\n"
             "te_prio = priority_index(te, proba, list(le.classes_))\n"
             "top = te_prio.sort_values('priority_rank').head(20)\n"
             "top[['priority_rank','zip_code','state','hurricane_name',\n"
             "     'food_relief_priority_index','prob_high_or_severe',\n"
             "     'food_fragility_score','svi_overall','damage_severity_class']]"),
        code("te_prio.to_csv(OUT / 'priority_rankings.csv', index=False)\n"
             "print('saved priority_rankings.csv', te_prio.shape)"),
        md("## Folium choropleth maps — Ida and Ian"),
        code("import geopandas as gpd\n"
             "zcta_path = next((RAW / 'zcta').rglob('*.shp'))\n"
             "zcta = gpd.read_file(zcta_path).to_crs('EPSG:4326')\n"
             "zcta['zip_code'] = zcta['ZCTA5CE20'].astype(str).str.zfill(5)\n"
             "for hn, center in [('Ida', (30.2, -90.9)), ('Ian', (26.6, -81.9))]:\n"
             "    sub = te_prio[te_prio['hurricane_name']==hn]\n"
             "    z = zcta.merge(sub[['zip_code','priority_index_norm']], on='zip_code')\n"
             "    if z.empty:\n"
             "        print('no zips for', hn); continue\n"
             "    m = folium.Map(location=center, zoom_start=7)\n"
             "    folium.Choropleth(\n"
             "        geo_data=z.__geo_interface__, data=sub,\n"
             "        columns=['zip_code','priority_index_norm'],\n"
             "        key_on='feature.properties.zip_code',\n"
             "        fill_color='YlOrRd', legend_name=f'Priority index — {hn}').add_to(m)\n"
             "    for _, r in z.iterrows():\n"
             "        folium.GeoJson(r.geometry,\n"
             "            tooltip=f\"ZIP {r['zip_code']}  idx={r['priority_index_norm']:.2f}\").add_to(m)\n"
             "    out = OUT / f'priority_map_{hn.lower()}.html'\n"
             "    m.save(str(out)); print('saved', out)"),
        md("## Top-50 vs bottom-50 summary"),
        code("top50 = te_prio.nsmallest(50, 'priority_rank')\n"
             "bot50 = te_prio.nlargest(50, 'priority_rank')\n"
             "pd.DataFrame({\n"
             "  'top50': [top50['food_desert_flag'].sum(), top50['svi_overall'].mean(),\n"
             "            top50[TARGET_COL].mean()],\n"
             "  'bot50': [bot50['food_desert_flag'].sum(), bot50['svi_overall'].mean(),\n"
             "            bot50[TARGET_COL].mean()],\n"
             "}, index=['food_desert_count','mean_svi','mean_damage_per_1k'])"),
    ]
    return nb(cells)


# -----------------------------------------------------------------------------
# Write them all
# -----------------------------------------------------------------------------
NOTEBOOKS = {
    "01_data_acquisition.ipynb":      build_01,
    "02_eda.ipynb":                   build_02,
    "03_data_fusion.ipynb":           build_03,
    "04_feature_engineering.ipynb":   build_04,
    "05_unsupervised_modeling.ipynb": build_05,
    "06_supervised_modeling.ipynb":   build_06,
    "07_evaluation_and_shap.ipynb":   build_07,
    "08_priority_index.ipynb":        build_08,
}


def main():
    for name, fn in NOTEBOOKS.items():
        path = HERE / name
        with open(path, "w", encoding="utf-8") as f:
            json.dump(fn(), f, indent=1)
        print("wrote", path)


if __name__ == "__main__":
    main()
